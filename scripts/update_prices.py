from openpyxl import load_workbook

from beevo.client import BeevoClient
from beevo.config.env_loader import load_environment
from beevo.variants import VariantsAPI


def load_prices_from_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = {}
    price_map = {}

    def normalize(text):
        return str(text).strip().lower()

    def normalize_price_header(text):
        header = normalize(text)
        replacements = {
            "Ã£Â§": "Ã§",
            "ÃƒÆ’Ã‚Â§": "Ã§",
            "ç": "Ã§",
            "preÃ£Â§o": "preÃ§o",
            "preÃƒÆ’Ã‚Â§o": "preÃ§o",
            "preço": "preÃ§o",
            "Ã¢â€šÂ¬": "â‚¬",
            "ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬": "â‚¬",
            "€": "â‚¬",
        }

        for source, target in replacements.items():
            header = header.replace(source, target)

        return header

    def parse_price(raw_value):
        if isinstance(raw_value, str):
            normalized_price = (
                raw_value
                .replace("ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬", "")
                .replace("Ã¢â€šÂ¬", "")
                .replace("â‚¬", "")
                .replace("€", "")
                .replace(",", ".")
                .strip()
            )

            if not normalized_price or normalized_price.startswith("#"):
                return None

            try:
                return float(normalized_price)
            except ValueError:
                return None

        if raw_value is None:
            return None

        try:
            return float(raw_value)
        except (TypeError, ValueError):
            return None

    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[normalize(cell.value)] = col_idx

    sku_col = None
    price_col = None

    for header, idx in headers.items():
        if header == "ref":
            sku_col = idx

        normalized_header = normalize_price_header(header)
        if "preÃ§o loja" in normalized_header and "iva" in normalized_header:
            price_col = idx

    if not sku_col or not price_col:
        raise ValueError(
            "Excel must contain 'REF' and 'PreÃ§o Loja c/ IVA 23% (â‚¬)' columns"
        )

    for row in ws.iter_rows(min_row=2):
        sku_cell = row[sku_col - 1].value
        price_cell = row[price_col - 1].value

        if not sku_cell:
            continue

        sku = str(sku_cell).strip().upper()
        price = parse_price(price_cell)

        if price is None:
            print(f"[SKIP] Invalid price for {sku}: {price_cell}")
            continue

        price_map[sku] = price

    return price_map


def update_prices_from_map(variants_api, price_map, variant_lookup):
    print("[STEP] Updating prices...")

    for sku, new_price in price_map.items():
        sku = sku.strip().upper()

        if sku not in variant_lookup:
            print(f"[SKIP] SKU not found: {sku}")
            continue

        variant = variant_lookup[sku]
        variant_id = variant["id"]
        current_price = variant["price"]
        new_price_cents = int(new_price * 100)

        if current_price == new_price_cents:
            print(f"[SKIP] No change for {sku}")
            continue

        try:
            variants_api.update_variant(
                variant_id=variant_id,
                price=new_price_cents,
            )
            print(f"[OK] {sku}: {current_price} -> {new_price_cents}")
        except Exception as exc:
            print(f"[ERROR] {sku}: {exc}")


def run_price_update(excel_file):
    settings = load_environment()
    print("Environment loaded. Starting data collection...")
    print(f"Using Beevo URL: {settings.beevo_url}")

    client = BeevoClient(
        settings.beevo_url,
        settings.beevo_cookie,
        timeout=settings.request_timeout,
    )
    variants_api = VariantsAPI(client)

    price_map = load_prices_from_excel(excel_file)
    variant_lookup = variants_api.build_variant_lookup()
    update_prices_from_map(variants_api, price_map, variant_lookup)


if __name__ == "__main__":
    run_price_update("ecolux.xlsx")
