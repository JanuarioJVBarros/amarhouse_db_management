from pathlib import Path
import argparse

from openpyxl import load_workbook

from utils import json_utils


PROFILE_CONFIGS = {
    "efapel": {
        "json_path": "output.json",
        "excel_path": "TabelA PREcos EFAPEL.xlsx",
        "code_column": "REFERENCIA",
        "price_column": "PRECO LOJA +20%",
        "target": "variants",
        "sku_prefix": "",
    },
    "golmar": {
        "json_path": "scraped_products copy.json",
        "excel_path": "Tabela_Golmar_PVP_2026.xlsx",
        "code_column": "CODIGO",
        "price_column": "preco loja+20",
        "target": "products",
        "sku_prefix": "G",
    },
}


def normalize_header(value):
    text = str(value or "").strip().lower()
    replacements = {
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "ä": "a",
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "í": "i",
        "ì": "i",
        "î": "i",
        "ï": "i",
        "ó": "o",
        "ò": "o",
        "ô": "o",
        "õ": "o",
        "ö": "o",
        "ú": "u",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "ç": "c",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return " ".join(text.split())


def load_price_dict(excel_path, code_column, price_column):
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    header_cells = [cell.value for cell in sheet[1]]
    normalized_columns = {
        normalize_header(column): index for index, column in enumerate(header_cells, start=1)
    }

    normalized_code_column = normalize_header(code_column)
    normalized_price_column = normalize_header(price_column)

    if normalized_code_column not in normalized_columns:
        raise ValueError(f"Missing code column: {code_column}")

    if normalized_price_column not in normalized_columns:
        raise ValueError(f"Missing price column: {price_column}")

    code_column_index = normalized_columns[normalized_code_column]
    price_column_index = normalized_columns[normalized_price_column]
    price_dict = {}

    for row in sheet.iter_rows(min_row=2):
        code = str(row[code_column_index - 1].value).strip()
        if not code or code == "nan":
            continue

        price = row[price_column_index - 1].value
        if price is None:
            continue

        price_dict[code] = float(price)

    return price_dict


def update_variant_prices(products, price_map):
    updated = 0
    missing = 0

    for product in products:
        for variant in product.get("variants", []):
            code = str(variant.get("sku") or "").strip()
            if not code:
                continue

            if code in price_map:
                variant["price"] = int(round(price_map[code], 2) * 100)
                updated += 1
            else:
                variant["price"] = 0
                missing += 1

    return updated, missing


def update_product_prices(products, price_map, sku_prefix=""):
    updated = 0
    missing = 0

    for product in products:
        sku = str(product.get("sku") or "").strip()
        code = f"{sku_prefix}{sku}".strip()
        if not code:
            continue

        if code in price_map:
            product["price"] = int(round(price_map[code], 2) * 100)
            updated += 1
        else:
            product["price"] = 0
            missing += 1

    return updated, missing


def sync_json_prices(
    json_path,
    excel_path,
    code_column,
    price_column,
    target="variants",
    sku_prefix="",
):
    json_file = Path(json_path)
    excel_file = Path(excel_path)

    print("Loading JSON...")
    products = json_utils.load_json(json_file)

    print("Loading Excel...")
    price_map = load_price_dict(excel_file, code_column, price_column)
    print(f"Loaded {len(price_map)} prices")

    print("Updating products...")
    if target == "products":
        updated, missing = update_product_prices(products, price_map, sku_prefix=sku_prefix)
    else:
        updated, missing = update_variant_prices(products, price_map)

    print(f"[OK] Updated: {updated}")
    print(f"[WARN] Missing in Excel: {missing}")

    print("Saving...")
    json_utils.save_json(json_file, products)
    print("[OK] Done")
    return products


def build_parser():
    parser = argparse.ArgumentParser(
        description="Sync scraped JSON prices from a supplier Excel sheet.",
    )
    parser.add_argument(
        "--profile",
        choices=sorted(PROFILE_CONFIGS),
        help="Optional supplier profile with sensible defaults.",
    )
    parser.add_argument("--json-file", help="JSON file to update.")
    parser.add_argument("--excel-file", help="Excel file to read prices from.")
    parser.add_argument("--code-column", help="Excel column that contains the SKU/code.")
    parser.add_argument("--price-column", help="Excel column that contains the price.")
    parser.add_argument(
        "--target",
        choices=("products", "variants"),
        help="Whether prices live on top-level products or product variants.",
    )
    parser.add_argument(
        "--sku-prefix",
        default=None,
        help="Optional prefix to prepend before matching JSON SKUs to Excel codes.",
    )
    return parser


def resolve_args(args):
    profile_config = PROFILE_CONFIGS.get(args.profile or "", {})
    resolved = {
        "json_path": args.json_file or profile_config.get("json_path"),
        "excel_path": args.excel_file or profile_config.get("excel_path"),
        "code_column": args.code_column or profile_config.get("code_column"),
        "price_column": args.price_column or profile_config.get("price_column"),
        "target": args.target or profile_config.get("target", "variants"),
        "sku_prefix": args.sku_prefix
        if args.sku_prefix is not None
        else profile_config.get("sku_prefix", ""),
    }

    missing = [key for key, value in resolved.items() if value is None]
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"Missing required arguments: {joined}")

    return resolved


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    resolved = resolve_args(args)
    return sync_json_prices(**resolved)


if __name__ == "__main__":
    main()
