from openpyxl import load_workbook

from scripts.find_missing_products_from_excel import (
    export_missing_to_excel,
    find_missing_skus,
)


def test_find_missing_skus_returns_only_missing_items():
    price_map = {
        "abc-1": 10.0,
        "missing-1": 20.5,
        "missing-2": 30.0,
    }
    variant_lookup = {
        "ABC-1": {"id": "variant-1", "price": 1000},
    }

    result = find_missing_skus(price_map, variant_lookup)

    assert result == [
        {"sku": "MISSING-1", "price": 20.5},
        {"sku": "MISSING-2", "price": 30.0},
    ]


def test_export_missing_to_excel_writes_rows(tmp_path):
    output_file = tmp_path / "missing_products.xlsx"
    export_missing_to_excel(
        [
            {"sku": "MISSING-1", "price": 20.5},
            {"sku": "MISSING-2", "price": 30.0},
        ],
        filename=output_file,
    )

    workbook = load_workbook(output_file)
    sheet = workbook.active

    assert sheet.title == "Missing Products"
    assert sheet.cell(row=1, column=1).value == "SKU"
    assert sheet.cell(row=1, column=2).value == "Price"
    assert sheet.cell(row=2, column=1).value == "MISSING-1"
    assert sheet.cell(row=2, column=2).value == 20.5
    assert sheet.cell(row=3, column=1).value == "MISSING-2"
    assert sheet.cell(row=3, column=2).value == 30.0
